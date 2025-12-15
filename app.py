import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import re
import io
import zipfile

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Gerador de Fichas Fast Shop", page_icon="🚚")

st.title("🚚 Gerador de Fichas de Fluxo")
st.markdown("Cole os dados brutos, faça upload do modelo em Excel e baixe tudo preenchido.")

# --- MAPEAMENTO DAS CÉLULAS (CONFIGURAÇÃO) ---
# Se o layout mudar, altere as células aqui (ColunaLinha)
MAPA_CELULAS = {
    'TRANSPORTADOR': 'B2',
    'CARGA': 'J2',
    'MOTORISTA': 'B4',
    'CNH': 'J4',
    'RG': 'B5',
    'CPF': 'J5',
    'TRUCK': 'B10',
    'CAVALO': 'B11',
    'CARRETA': 'B12'
}

# --- FUNÇÃO DE PROCESSAMENTO DE TEXTO ---
def processar_texto(texto_bruto):
    blocos = texto_bruto.strip().split('\n\n') # Separa por linha vazia dupla
    lista_dados = []

    for bloco in blocos:
        if len(bloco.strip()) < 10: continue # Pula sujeira ou blocos vazios
        
        linhas = bloco.strip().split('\n')
        dados = {}
        
        # Pega Transportadora e Carga (Assumindo linhas 1 e 2)
        if len(linhas) >= 2:
            dados['TRANSPORTADOR'] = linhas[0].strip()
            dados['CARGA'] = linhas[1].strip()
        
        texto_bloco = "\n".join(linhas)
        
        # Regex para capturar campos
        def extrair(padrao, texto):
            match = re.search(padrao, texto)
            return match.group(1).strip() if match else ""

        dados['MOTORISTA'] = extrair(r'MOT:\s*(.*)', texto_bloco)
        dados['CPF'] = extrair(r'CPF:\s*([\d.-]+)', texto_bloco)
        dados['RG'] = extrair(r'RG:\s*([\d]+)', texto_bloco)
        dados['CNH'] = extrair(r'CNH:\s*([\d]+)', texto_bloco)
        
        # Lógica Truck vs Conjunto
        truck = extrair(r'TRUCK:\s*([A-Z0-9]+)', texto_bloco)
        cavalo = extrair(r'CAVALO:\s*([A-Z0-9]+)', texto_bloco)
        carreta = extrair(r'CARRETA:\s*([A-Z0-9]+)', texto_bloco)

        if truck:
            dados['TRUCK'] = truck
            dados['CAVALO'] = ""
            dados['CARRETA'] = ""
        else:
            dados['TRUCK'] = ""
            dados['CAVALO'] = cavalo
            dados['CARRETA'] = carreta
            
        lista_dados.append(dados)
    return lista_dados

# --- INTERFACE DO USUÁRIO ---

# 1. Upload do Modelo
st.subheader("1. Faça upload do Modelo em Branco (.xlsx)")
arquivo_modelo = st.file_uploader("Arraste seu modelo Excel aqui", type=["xlsx"])

# 2. Área de Texto
st.subheader("2. Cole os dados dos motoristas")
texto_input = st.text_area("Cole aqui (CARRARO, SC+RS...)", height=200)

# 3. Botão de Processar
if st.button("Gerar Fichas", type="primary"):
    if not arquivo_modelo:
        st.error("Por favor, faça o upload do arquivo modelo primeiro.")
    elif not texto_input:
        st.error("Por favor, cole os dados dos motoristas.")
    else:
        # Processamento
        try:
            dados_processados = processar_texto(texto_input)
            
            # Criar arquivo ZIP em memória
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, "w") as zip_file:
                for i, registro in enumerate(dados_processados):
                    # Carrega o modelo da memória
                    arquivo_modelo.seek(0)
                    wb = load_workbook(arquivo_modelo)
                    ws = wb.active
                    
                    # Preenche
                    for campo, valor in registro.items():
                        celula = MAPA_CELULAS.get(campo)
                        if celula:
                            ws[celula] = valor
                    
                    # Salva o Excel individual na memória
                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    nome_arquivo = f"Fluxo_{i+1}_{registro['MOTORISTA'][:10].replace(' ','_')}.xlsx"
                    
                    # Adiciona ao ZIP
                    zip_file.writestr(nome_arquivo, excel_buffer.getvalue())

            # Botão de Download do ZIP
            st.success(f"{len(dados_processados)} fichas geradas com sucesso!")
            st.download_button(
                label="📥 Baixar Todas as Fichas (ZIP)",
                data=zip_buffer.getvalue(),
                file_name="Fichas_Preenchidas.zip",
                mime="application/zip"
            )
            
            # Mostra prévia na tela para conferência
            st.write("---")
            st.subheader("Prévia dos dados identificados:")
            st.dataframe(pd.DataFrame(dados_processados))

        except Exception as e:
            st.error(f"Ocorreu um erro: {e}")